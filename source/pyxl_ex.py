from openpyxl import Workbook

wb = load_workbook('test.xlsx')
ws = wb.active

ws['A4'] = "Cell A4"
a = ws['A4']

b = ws.cell(row = 4, column = 2, value = "B4")

print(a.value)
print(b.value)

for row in ws.values:
    for value in row:
        print(value)


wb.save(filename='test.xlsx')