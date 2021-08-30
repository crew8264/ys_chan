from openpyxl import load_workbook, Workbook

c_name = 1
c_id = 2
c_lvl = 3
c_exp = 4
c_money = 5
c_loss = 6

def checkName(_name, _id):
    wb = load_workbook('/YS_CHAN/etc./userDB.xlsx')
    ws = wb.active

    for row in range(2, ws.max_row+2):
        if ws.cell(row, c_name).value == _name and ws.cell(row.c_id).value == hex(_id):
            wb.save("userDB.xlsx")
            wb.close()
            return False
            break
        else:
            print("research : ")
        
    wb.save("/YS_CHAN/etc./userDB.xlsx")
    wb.close()
